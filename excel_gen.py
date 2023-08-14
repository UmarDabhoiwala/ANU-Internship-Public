import requests
import pandas as pd 
from io import StringIO
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Series, Reference, PieChart


def ExcelGen(numberOfRows, decision, graphsWanted = False):
    
    if numberOfRows > 1000:
        numberOfRows = 1000
    
    if numberOfRows <= 0:
        numberOfRows = 1
    
    
    params = {
        'count': numberOfRows,
        'key': '4cdfcc50',
    }

    defaultAPI = "1f1a6ed0"
    customerDataAPI = "54294940"
    investmentPortfolioAPI = "3c5fec20"
    programBudgeting = "1d437da0"
    accountBalances = "dc8bb490"

    if decision == 1:
        api = defaultAPI
        nameA = "Employee"
        nameB = "Data"
    elif decision == 2:
        api = customerDataAPI
        nameA = "Customer"
        nameB = "Data"
    elif decision == 3:
        api = investmentPortfolioAPI
        nameA = "Investment"
        nameB = "Portfolio"
    elif decision == 4:
        api = programBudgeting
        nameA = "Program"
        nameB = "Budgeting"
    elif decision == 5:
        api = accountBalances
        nameA = "Account"
        nameB = "Balances"
    else:
        api = defaultAPI
        nameA = "Employee"
        nameB = "Data"
      
            
    response = requests.get(f'https://api.mockaroo.com/api/{api}', params=params)
    
    print(response.status_code)

    data = StringIO(response.text)
    df = pd.read_csv(data)
    
    sheetName = nameA + " " + nameB


    df.to_excel("files/base.xlsx", sheet_name=sheetName, engine = "xlsxwriter", startrow= 3, index= False)

    wb = load_workbook('files/base.xlsx')
    ws = wb.active


    #Widens columns to the longest entry 
    def columnWidener (sheet):
        dims = {}
        for i, row in enumerate (sheet.rows):
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
            
            #Don't want to check every row too much computation
            if i > 100:
                break
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value +5
            
    columnWidener(ws)

    def rowHighlighter (sheet, minRow, maxRow, colour = 'FFFF00'):

        for row in sheet.iter_rows(min_row=minRow, max_row=maxRow):
            # Check if any cell in the row contains a value
            if any(cell.value is not None for cell in row):
                # If so, set the fill color for all cells in the row
                fill = PatternFill(start_color=colour, end_color=colour, fill_type='solid')
                for cell in row:
                    cell.fill = fill

    rowHighlighter(ws, 4, 4)

    orangeFill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type='solid')
    ws["A2"] = nameA
    ws["A2"].font = Font(size=16)
    ws["A2"].fill = orangeFill
    ws["B2"] = nameB
    ws["B2"].font = Font(size=16)
    ws["B2"].fill = orangeFill



    columnWidener(ws)


    def uniqueCountGraph (title, column, columnName, nameOfNewSheet, YTitle, XTitle, location, width = 15, height = 10): 
    # Create a new workbook and worksheet to hold the graph data

        # Iterate through the column of interest and count the number of occurrences of each unique value
        value_counts = {}
        for cell in ws[column]:
            if cell.value not in value_counts and cell.value not in (None, columnName):
                value_counts[cell.value] = 1
            else:
                if cell.value not in (None, columnName):
                    value_counts[cell.value] += 1



        wb.create_sheet(nameOfNewSheet)
        chart_sheet = wb[nameOfNewSheet]

        # Write the unique values and their counts to the worksheet
        chart_sheet.cell(row=1, column =1, value = YTitle)
        chart_sheet.cell(row=1, column =2, value = XTitle)
        for i, (value, count) in enumerate(value_counts.items(), start=2):
            chart_sheet.cell(row=i, column=1, value=value)
            chart_sheet.cell(row=i, column=2, value=count)
        
        rowHighlighter(chart_sheet, 1, 1)
        columnWidener(chart_sheet)
        

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = YTitle
        chart.x_axis.title = XTitle

        dataCol = Reference(chart_sheet, min_col=2, max_col =2, min_row=1, max_row= len (value_counts) + 1)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row = len (value_counts) + 1)

        chart.add_data(data = dataCol, titles_from_data = True)
        chart.set_categories(cats)
        chart.shape = 4
        
        chart.width = width
        chart.height = height

        # Add the chart to the original worksheet
        ws.add_chart(chart, location)
        
        
    def piChart(sheet, title, dataCol, labelCol, loc, width = 15, height = 10):
        
        pie = PieChart()
        
        labels = Reference(sheet, min_col=labelCol, min_row=5, max_row= sheet.max_row)
        data_ref = Reference(sheet, min_col=dataCol, min_row=4, max_row= sheet.max_row)
    
        
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = title
        
        pie.width = width
        pie.height = height
        
        sheet.add_chart(pie, loc)

            
    if graphsWanted:
        if decision == 1:
            uniqueCountGraph("Gender Comparison","E", "gender", "genderData", "Number","Gender","J4")
            uniqueCountGraph("Department Head Count", "F", "department", "departmentData", "Number of Employees", "Department", "J30", 20, 15)
        if decision == 3: 
            piChart(ws, "Total Holding Distribution", 4, 6, "I4", width = 30, height=20)
            piChart(ws, "Shares Distribution", 3, 6, "I50", width = 30, height=20)
            

    # Save the workbook with the chart
    wb.save('files/base.xlsx')
    

